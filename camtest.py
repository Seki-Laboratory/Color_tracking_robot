import os
import glob
import imghdr
import openpyxl
import cv2
 
 
# 定数設定
INPUT_IMG_DIR = 'C:\\Users\\usui0\\Pictures\\Camera Roll\\1' # 貼り付ける画像を置いておくルートディレクトリ
SHEET_TITLE = '画像貼り付け' # シート名の設定
RESULT_FILE_NAME = '.\\result.xlsx' # 結果を保存するファイル名
 
# 変数
max_height = [] # 各行の画像の高さの最大値を保持
 
 
def get_file_names(set_dir_name):
    """
    ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
    """
    file_names = os.listdir(set_dir_name)
    temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(os.path.join(set_dir_name, file_name))] # ファイルかどうかを判定
    return temp_full_file_names
 
def attach_img(target_full_file_names, set_column_idx, set_dir_name):
    """
    画像を呼び出して、Excelに貼り付け
    """
    set_row_idx = 1
    column_letter = ws.cell(row=set_row_idx, column=set_column_idx).column # セルの行列番号から、そのセルの列番号の文字列を取得
    ws.cell(row=1, column=set_column_idx).value = set_dir_name # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
    max_width = 0 # 画像の幅の最大値を保持するための変数
    target_full_file_names.sort() # ファイル名でソート
    for target_file in target_full_file_names:
        if imghdr.what(target_file) != None: # 画像ファイルかどうかの判定
            img = openpyxl.drawing.image.Image(target_file)
            #print('[' + str(column_letter) + '][' + str(set_row_idx+1) + ']' + str(target_file) + 'を貼り付け')
 
            # 画像のサイズを取得して、セルの大きさを合わせる（画像同士が重ならないようにするため）
            size_img = cv2.imread(target_file)
            height, width = size_img.shape[:2]
            if max_width < width:
                max_width = width
            if not max_height[set_row_idx-1:set_row_idx]: # 配列「max_height」において、「set_row_idx」番目の要素が存在しなければ、挿入
                max_height.insert(set_row_idx-1, height)
            if max_height[set_row_idx-1] < height:
                max_height[set_row_idx-1] = height
            ws.row_dimensions[set_row_idx+1].height = max_height[set_row_idx-1] * 0.75
            ws.column_dimensions[column_letter].width = max_width * 0.13
 
            cell_address = ws.cell(row=set_row_idx + 1, column=set_column_idx).coordinate # セルの行列番号から、そのセルの番地を取得
            img.anchor = cell_address
            ws.add_image(img) # シートに画像貼り付け
 
        set_row_idx += 1
 
 
# ワークブック設定
wb = openpyxl.Workbook()
ws = wb.worksheets[0] # 1番目のシートを編集対象にする
ws.title = SHEET_TITLE # 1番目のシートに名前を設定
 
# 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
dirs = glob.glob(os.path.join(INPUT_IMG_DIR, '**' + os.sep), recursive=True)
 
column_idx = 1
# 各ディレクトリについて操作
for dir_name in dirs:
    f_names = get_file_names(dir_name) # ファイル名取得
    attach_img(f_names, column_idx, dir_name) # 画像貼り付け設定
    column_idx += 1 # 次の列へ・・・
 
# ファイルへの書き込み
wb.save(RESULT_FILE_NAME)